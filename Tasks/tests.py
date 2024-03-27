import subprocess
import time
import openpyxl
import pandas as pd
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from faker import Faker
from Utilities import drivers
from bs4 import BeautifulSoup
from Utilities.config import FILEPATH

fake = Faker()
driver = drivers.initialize_driver("chrome")

driver.get("https://www.saucedemo.com/v1/")
driver.maximize_window()


# Web_Elements

username_element = driver.find_element(By.ID, 'user-name')
password_element = driver.find_element(By.ID, 'password')
login_btn_element = driver.find_element(By.ID, 'login-button')


def get_user_credentials():
    login_credentials_div = driver.find_element(By.CLASS_NAME, "login_credentials_wrap")
    login_credentials_div_html = login_credentials_div.get_attribute('innerHTML')
    soup = BeautifulSoup(login_credentials_div_html, 'html.parser')

    login_usernames = soup.find('div', id='login_credentials')

    if login_usernames:
        br_tags = login_usernames.find_all('br')
        usernames = [br_tag.previous_sibling.strip() for br_tag in br_tags if br_tag.previous_sibling]

        user_id = []
        for user in usernames:
            user_id.append(fake.uuid4()[:4])
        df_credentials = pd.DataFrame({'User ID': user_id, "Username": usernames, "Password": "secret_sauce"})
        df_credentials.to_excel('../Test_Data/user_credentials.xlsx', index=False)

# def run_tests():
#     subprocess.run(['python', 'tests.py'])
def login_and_record_errors():
    df = pd.read_excel(FILEPATH, sheet_name="Sheet1")

    wb = openpyxl.load_workbook(FILEPATH)
    if "Login" not in wb.sheetnames:
        ws = wb.create_sheet("Login")
    else:
        ws = wb["Login"]
    ws.append(["User ID", "Login Message"])

    for index, row in df.iterrows():
        user_id = row['User ID']
        username = row['Username']
        password = row['Password']

        # time.sleep(2)
        driver.find_element(By.ID, 'user-name').send_keys(username)
        driver.find_element(By.ID, 'password').send_keys(password)
        driver.find_element(By.ID, 'login-button').click()

        try:
            # time.sleep(2)
            error_message_element = driver.find_element(By.CSS_SELECTOR, "h3[data-test='error']")
            error_message = error_message_element.text
            ws.append([user_id, error_message])
        except NoSuchElementException:
            # time.sleep(2)
            ws.append([user_id, "Logged in Successfully"])

        driver.get("https://www.saucedemo.com/v1/")
        # time.sleep(2)

    wb.save(FILEPATH)


def standard_user_product_details():
    username_element_standard = driver.find_element(By.ID, "user-name")
    password_element_standard = driver.find_element(By.ID, "password")
    login_btn_element_standard = driver.find_element(By.ID, "login-button")
    username = "standard_user"
    password = "secret_sauce"

    username_element_standard.send_keys(username)
    password_element_standard.send_keys(password)
    login_btn_element_standard.click()

    products = driver.find_elements(By.CLASS_NAME, "inventory_item")
    product_details = {"Product ID": [], "Product Name": [], "Description": [], "Price": []}

    for product in products:
        product_info = product.text.split("\n")
        product_name = product_info[0]
        product_description = product_info[1]
        product_price = product_info[2]

        product_details["Product ID"].append(fake.uuid4()[:4])
        product_details["Product Name"].append(product_name)
        product_details["Description"].append(product_description)
        product_details["Price"].append(product_price)

    df = pd.DataFrame(product_details)
    with pd.ExcelWriter(FILEPATH, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='Products details', index=False)


def orders():
    user_product_data = pd.ExcelFile(FILEPATH)
    sheet_names = user_product_data.sheet_names

    if "Order Details" not in sheet_names:
        orders_df = pd.DataFrame(
            columns=['Order ID', 'User ID', 'Product ID', 'Product Name', 'Quantity', 'Total Price'])
    else:
        orders_df = user_product_data.parse('Order Details')

    standard_user_orders = [
        {'Order ID': '1', 'User ID': 'standard_user', 'Product ID': '1', 'Product Name': 'Sauce Labs Backpack',
         'Quantity': 2,
         'Total Price': 59.98},
        {'Order ID': '2', 'User ID': 'standard_user', 'Product ID': '3', 'Product Name': 'Sauce Labs Bike Light',
         'Quantity': 1,
         'Total Price': 9.99}, ]

    problem_user_orders = [
        {'Order ID': '3', 'User ID': 'problem_user', 'Product ID': '2', 'Product Name': 'Sauce Labs Backpack',
         'Quantity': 1,
         'Total Price': 29.99}, ]

    orders_df = pd.concat([orders_df, pd.DataFrame(standard_user_orders + problem_user_orders)])

    with pd.ExcelWriter(FILEPATH, mode='a', engine='openpyxl', if_sheet_exists='replace') as orders_writer:
        orders_df.to_excel(orders_writer, sheet_name='Order Details', index=False)





