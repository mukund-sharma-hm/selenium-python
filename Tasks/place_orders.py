"""
python
"""
import sys
import logging
import time
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException
from selenium.common import NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

from Utilities import drivers
from Utilities.config import FILEPATH, PASSWORD

sys.path.append('../Utilities')
sys.path.append('../Drivers')


class PlaceOrders:
    """
    Class to manage placing orders based on data from an Excel file.

    Attributes:
    wait (WebDriverWait): The WebDriverWait object.
    driver (WebDriver): The WebDriver object.
    excel (str): Path to the Excel file.
    wb (Workbook): The Workbook object.
    """

    def __init__(self):
        """
        Constructor for the PlaceOrders class.
        Initializes attributes for driver, wait, excel, and workbook.
        """
        # self.driver = None
        # Initialize attributes
        self.wait: Optional[WebDriverWait] = None
        self.driver: Optional[WebDriver] = None
        self.excel = None
        self.wb: Optional[Workbook] = None

        # initialize logger
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.DEBUG)

        # handle in console
        # self.handler = logging.StreamHandler()
        # self.handler.setLevel(logging.DEBUG)
        # self.logger.addHandler(self.handler)

        # Create a file handler for logging to 'logs.txt' file
        self.file_handler = logging.FileHandler("../Logs/logs.txt")
        self.file_handler.setLevel(logging.DEBUG)

        # Create a formatter for the file handler
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        self.file_handler.setFormatter(formatter)

        # Add the file handler to the logger
        self.logger.addHandler(self.file_handler)

    def initialize_driver(self):

        """
        Initializes the WebDriver and WebDriverWait for the driver.
        """
        try:
            self.driver = drivers.initialize_driver("chrome")
            self.logger.info("Driver initialized")
            self.wait = WebDriverWait(self.driver, 2)
        except WebDriverException:
            self.logger.error("Could not initialize driver")


    def load_excel(self, excel_file):
        """
        Load an Excel file and store the workbook for further processing.

        Parameters:
        excel_file (str): The path to the Excel file to load.

        Returns:
        None
        """
        try:
            self.excel = excel_file
            self.wb = openpyxl.load_workbook(FILEPATH)
            self.logger.info("Excel loaded successfully")
        except InvalidFileException:
            self.logger.error("File can not be loaded, invalid format")
    def login_user(self, username, password):
        """
        Log in a user with the provided username and password.

        Parameters:
        username (str): The username to log in with.
        password (str): The password for the user.

        Returns:
        None
        """
        try:
            self.driver.get("https://www.saucedemo.com/v1/")
            self.driver.maximize_window()
            self.driver.find_element(By.ID, "user-name").send_keys(username)
            self.driver.find_element(By.ID, "password").send_keys(password)
            self.driver.find_element(By.ID, "login-button").click()
            self.logger.info("Logged in successfully as %s", username)
        except NoSuchElementException:
            self.logger.error("Login failed, element could not be found")
        except TimeoutException:
            self.logger.error("Login failed, Timed out finding the element")

    def place_orders(self):
        """
        Place orders based on the data loaded from the Excel file.

        Parameters:
        None

        Returns:
        None
        """

        self.logger.info("starting to place orders")
        order_details = self.wb['Order Details']
        sheet = self.wb["Order Details"]

        for row in order_details.iter_rows(min_row=2, values_only=True):
            quantity_in_site = 0
            # Extract data from the row
            if len(row) == 6:
                self.logger.debug("Process row %s", row)
                order_id, user_id, product_id, product_name, quantity, total_price, = row
            else:
                (order_id, user_id, product_id, product_name,
                 quantity, total_price, order_status) = row

            # Login user with provided user_id and PASSWORD
            self.logger.info("Login user id: %s", user_id)
            self.login_user(user_id, PASSWORD)

            # Find and click on the product
            self.logger.info("clicking on product: %s", product_name)
            product_name_element = self.driver.find_element(By.XPATH,
                                                            f"//*[text()='{product_name}']")
            product_name_element.click()
            self.logger.info("product viewed")
            # Add the product to the cart
            add_to_cart_element = self.driver.find_element(By.XPATH, "//*[text()='ADD TO CART']")
            add_to_cart_element.click()
            self.logger.info("product: %s added to the cart", product_name)

            # Proceed to the cart and checkout
            cart_element = self.driver.find_element(By.ID, 'shopping_cart_container')
            cart_element.click()
            time.sleep(5)
            try:
                quantity_in_site = self.driver.find_element(By.XPATH, "//*[@class='cart_quantity']").text
            except NoSuchElementException:
                self.logger.warning("Quantity of the product not found")
                order_status = "failure"
            checkout_element = self.driver.find_element(By.XPATH, "//*[@class='cart_footer']/a[2]")
            checkout_element.click()

            # Fill in checkout details
            self.driver.find_element(By.ID, "first-name").send_keys("random")
            self.driver.find_element(By.ID, "last-name").send_keys("lastname")
            self.driver.find_element(By.ID, "postal-code").send_keys("175101")
            self.driver.find_element(By.XPATH, "//*[@class='checkout_buttons']/input").click()

            try:
                # Check if the total price matches and set order status
                total_item_price_element = self.driver.find_element(By.XPATH, "//*[@class='summary_subtotal_label']")
                total_item_price_element_text = total_item_price_element.text
                total_item_price = total_item_price_element_text.split("$")[1]
                total_price = total_price/quantity
                if str(total_price) == total_item_price and quantity_in_site == '1':
                    order_status = "success"
                else:
                    order_status = "failure"
                self.driver.find_element(By.XPATH, "//*[text()='FINISH']").click()

            except NoSuchElementException:
                self.logger.warning("Element for finish button not found")
                order_status = "failure"

            if order_status == "success":
                try:
                    # Check for success message and update order status
                    success_message = self.wait.until(
                        EC.visibility_of_element_located((By.XPATH,
                                                          "//h2[text()='THANK YOU FOR YOUR ORDER']")))

                    if success_message.is_displayed():
                        order_status = "success"

                except TimeoutException:
                    self.logger.error("Timed out")
                    order_status = "Failure"

            # Update the order status in the Excel sheet
            column_index = 7
            sheet.cell(row=1, column=column_index, value="Order Status")
            sheet.cell(row=int(order_id) + 1, column=column_index, value=order_status)
            self.wb.save(FILEPATH)
        self.logger.info("order placement completed")
